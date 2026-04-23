import asyncio
import smtplib
import sys
import os
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()  # load .env without overriding already-set env vars (e.g. from GitHub Actions)

# Ensure required vars are set before importing intervals_mcp_server (config reads them at import time)
if not os.environ.get("ATHLETE_ID"):
    os.environ["ATHLETE_ID"] = "i418663"
if not os.environ.get("INTERVALS_API_BASE_URL"):
    os.environ["INTERVALS_API_BASE_URL"] = "https://intervals.icu/api/v1"

sys.path.insert(0, str(Path(__file__).parent / "src"))

import intervals_mcp_server.config as cfg
cfg._config_instance = None

from intervals_mcp_server.tools.activities import get_activities

OUTPUT_PATH = Path(__file__).parent / "weekly_running_plan.xlsx"
SEND_TO = "miki31877@gmail.com"
GMAIL_USER = "miki31877@gmail.com"
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")


async def fetch_runs():
    today = datetime.today().strftime("%Y-%m-%d")
    start = (datetime.today() - timedelta(days=120)).strftime("%Y-%m-%d")
    raw = await get_activities(start_date=start, end_date=today, limit=50, include_unnamed=False)
    print("API response preview:", raw[:300])

    runs = []
    for block in raw.split("\n\nActivity:"):
        lines = {}
        for line in block.splitlines():
            if ": " in line:
                k, _, v = line.partition(": ")
                lines[k.strip()] = v.strip()
        if lines.get("Type") == "Run":
            try:
                dist = float(lines.get("Distance", "0").replace(" meters", "") or 0)
                hr = float(lines.get("Average Heart Rate", "0").replace(" bpm", "") or 0)
                ctl = float(lines.get("Fitness (CTL)", "0") or 0)
                atl = float(lines.get("Fatigue (ATL)", "0") or 0)
                speed = float(lines.get("Average Speed", "0").replace(" m/s", "") or 0)
                date_str = lines.get("Date", "")[:10]
                runs.append({"dist": dist, "hr": hr, "ctl": ctl, "atl": atl,
                             "speed": speed, "date": date_str})
            except (ValueError, AttributeError):
                continue
    return runs


def analyze(runs):
    if not runs:
        return {"ctl": 25, "atl": 33, "avg_weekly_km": 20, "avg_hr": 156,
                "longest_km": 7.1, "pace_min_km": 6.4}

    latest = runs[0]
    ctl = latest["ctl"]
    atl = latest["atl"]

    # Average weekly km over last 4 weeks
    cutoff = (datetime.today() - timedelta(weeks=4)).strftime("%Y-%m-%d")
    recent = [r for r in runs if r["date"] >= cutoff]
    total_km = sum(r["dist"] for r in recent) / 1000
    avg_weekly_km = round(total_km / 4, 1)

    avg_hr = round(sum(r["hr"] for r in runs if r["hr"] > 0) / max(len(runs), 1), 0)
    longest_km = round(max(r["dist"] for r in runs) / 1000, 1)
    speeds = [r["speed"] for r in runs if r["speed"] > 0]
    avg_speed = sum(speeds) / len(speeds) if speeds else 2.6
    pace_min_km = round(1000 / avg_speed / 60, 1) if avg_speed > 0 else 6.5

    return {"ctl": round(ctl, 1), "atl": round(atl, 1),
            "avg_weekly_km": avg_weekly_km, "avg_hr": int(avg_hr),
            "longest_km": longest_km, "pace_min_km": pace_min_km}


def generate_excel(stats):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    rest_fill    = PatternFill("solid", fgColor="D9D9D9")
    easy_fill    = PatternFill("solid", fgColor="C6EFCE")
    quality_fill = PatternFill("solid", fgColor="FFEB9C")
    long_fill    = PatternFill("solid", fgColor="BDD7EE")
    white_fill   = PatternFill("solid", fgColor="FFFFFF")
    value_fill   = PatternFill("solid", fgColor="EBF3FB")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font   = Font(bold=True, size=10)
    normal_font = Font(size=10)
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(cell, fill, font, halign="center", wrap=True):
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
        cell.border = border

    # ── Sheet 1: Weekly Plan ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Weekly Plan"
    cols = ["Day", "Session", "Target HR", "Distance", "Pace Guide", "Details"]
    widths = [14, 22, 18, 16, 20, 52]

    for i, (col, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=1, column=i, value=col)
        cell_style(c, header_fill, header_font)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    long_dist = min(round(stats["longest_km"] + 1, 0), 14)
    long_dist_str = f"{int(long_dist)} km"

    plan = [
        ("Sunday",    "Long Easy Run", "< 145 bpm",   long_dist_str, "~7:00–7:45 /km",
         f"Most important session. Keep fully easy throughout. Target {long_dist_str} this week, adding ~1 km/week toward 14 km."),
        ("Monday",    "Rest / Walk",   "—", "—", "—", "Active recovery after the long run."),
        ("Tuesday",   "Rest",          "—", "—", "—", "Busy day — full rest."),
        ("Wednesday", "Easy Run",      "< 145 bpm",   "5–6 km",      "~7:00–7:30 /km",
         "Conversational pace. Slow down if HR creeps above 145."),
        ("Thursday",  "Quality Run",   "155–163 bpm", "5–6 km",      "~6:00–6:30 /km",
         "Easy 10 min warm-up, 15–20 min at threshold effort, easy 10 min cool-down."),
        ("Friday",    "Rest",          "—", "—", "—", "Full rest before the long run."),
        ("Saturday",  "Rest",          "—", "—", "—", "Full rest."),
    ]
    fills = [long_fill, rest_fill, rest_fill, easy_fill, quality_fill, rest_fill, rest_fill]

    for row, (data, fill) in enumerate(zip(plan, fills), 2):
        ws.row_dimensions[row].height = 42
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            cell_style(c, fill, bold_font if col == 1 else normal_font,
                       halign="center" if col < 5 else "left")

    # ── Sheet 2: Analysis ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Analysis")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 55
    section_fill = PatternFill("solid", fgColor="1F4E79")

    def add_section(row, title):
        c = ws2.cell(row=row, column=1, value=title)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell_style(c, section_fill, Font(bold=True, color="FFFFFF", size=11), halign="left")
        ws2.row_dimensions[row].height = 20
        return row + 1

    def add_row(row, label, value, highlight=False):
        lc = ws2.cell(row=row, column=1, value=label)
        vc = ws2.cell(row=row, column=2, value=value)
        f = value_fill if highlight else white_fill
        cell_style(lc, f, Font(bold=True, size=10), halign="left")
        cell_style(vc, f, Font(size=10), halign="left")
        ws2.row_dimensions[row].height = 28
        return row + 1

    over_threshold = stats["avg_hr"] >= 150
    observation = (
        "Running too hard on easy days — avg HR close to LTHR (163 bpm) on most runs."
        if over_threshold else
        "Good aerobic base — avg HR is comfortably below LTHR."
    )

    r = 1
    r = add_section(r, "Current Fitness Status")
    r = add_row(r, "Fitness (CTL)", stats["ctl"], True)
    r = add_row(r, "Fatigue (ATL)", stats["atl"])
    r = add_row(r, "Avg weekly km (last 4 wks)", f"{stats['avg_weekly_km']} km", True)
    r = add_row(r, "Longest recent run", f"{stats['longest_km']} km")
    r = add_row(r, "Avg pace on runs", f"{stats['pace_min_km']} min/km", True)

    r += 1
    r = add_section(r, "Heart Rate Analysis")
    r = add_row(r, "LTHR", "163 bpm")
    r = add_row(r, "Avg HR on runs", f"{stats['avg_hr']} bpm", True)
    r = add_row(r, "Key observation", observation)

    r += 1
    r = add_section(r, "8-Week Progression Target")
    r = add_row(r, "Sunday long run now", f"{int(long_dist)} km", True)
    r = add_row(r, "Target in 8 weeks", f"{min(int(long_dist) + 8, 14)} km")
    r = add_row(r, "Weekly volume target", "25–30 km/week (Sun long + Wed easy + Thu quality)", True)

    wb.save(OUTPUT_PATH)
    print(f"Excel saved: {OUTPUT_PATH}")
    return OUTPUT_PATH


def send_email(xlsx_path):
    if not GMAIL_APP_PASSWORD:
        print("GMAIL_APP_PASSWORD not set in .env — skipping email.")
        return

    msg = MIMEMultipart()
    msg["From"] = GMAIL_USER
    msg["To"] = SEND_TO
    msg["Subject"] = f"Weekly Running Plan - {datetime.today().strftime('%Y-%m-%d')}"
    msg.attach(MIMEText(
        "Hi,\n\nPlease find attached your updated weekly running plan "
        "generated from your latest Intervals.icu data.\n\nGood luck this week!",
        "plain"
    ))

    with open(xlsx_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{xlsx_path.name}"')
    msg.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_USER, SEND_TO, msg.as_string())
    print(f"Email sent to {SEND_TO}")


async def main():
    print("Fetching runs from Intervals.icu...")
    runs = await fetch_runs()
    print(f"Found {len(runs)} runs.")
    stats = analyze(runs)
    print(f"Stats: {stats}")
    xlsx = generate_excel(stats)
    send_email(xlsx)


if __name__ == "__main__":
    asyncio.run(main())
